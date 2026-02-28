"""
Clean and polish math dataset using Qwen2.5-72B-Instruct.

Supports both .csv and .xlsx input files.

For each row, the 72B model:
  1. Polishes the question (fix formatting, LaTeX, clarity)
  2. Polishes the answer (clean formatting, ensure proper LaTeX)
  3. Verifies the answer is correct by solving the problem

Outputs:
  - math_cleaned.csv        (cleaned dataset)
  - math_clean_log.jsonl    (detailed log with original/cleaned/verification)

Usage:
  python clean_math_csv.py \
    --model Qwen/Qwen2.5-72B-Instruct \
    --input ../raw/math.xlsx \
    --output_dir ../cleaned \
    --tp 4 --batch_size 32
"""

import argparse
import csv
import json
import os
import re
import time

from vllm import LLM, SamplingParams
from transformers import AutoTokenizer


CLEAN_SYSTEM_PROMPT = """\
You are a math dataset curator. Your job is to clean and polish math problems and their answers.

You will receive a math problem (question) and its reference answer. You must:

1. **Polish the question**:
   - Ensure all LaTeX expressions are properly formatted with $...$ delimiters.
   - Remove stray non-breaking spaces (\xa0) or other invisible characters.
   - Make the problem statement clear, complete, and grammatically correct.
   - Preserve the original language (Chinese or English). Do NOT translate.
   - If the question is already clean, keep it as-is.

2. **Polish the answer**:
   - Ensure proper LaTeX formatting.
   - Simplify if possible while preserving mathematical correctness.
   - Use consistent notation.

3. **Verify the answer**:
   - Solve the problem step by step.
   - Determine if the reference answer is correct.
   - If incorrect, provide the correct answer.

You MUST output in EXACTLY this JSON format (no other text):
{
  "question": "<cleaned question>",
  "answer": "<cleaned/corrected answer>",
  "is_correct": true/false,
  "corrected_answer": "<correct answer if is_correct is false, otherwise same as answer>",
  "verification_note": "<brief explanation of your verification>"
}"""

CLEAN_USER_TEMPLATE = """Question: {question}

Reference Answer: {answer}"""


def build_clean_prompts(questions, answers, tokenizer):
    """Build cleaning/verification prompts."""
    prompts = []
    for q, a in zip(questions, answers):
        messages = [
            {"role": "system", "content": CLEAN_SYSTEM_PROMPT},
            {"role": "user", "content": CLEAN_USER_TEMPLATE.format(question=q, answer=a)},
        ]
        prompt = tokenizer.apply_chat_template(
            messages, tokenize=False, add_generation_prompt=True
        )
        prompts.append(prompt)
    return prompts


def parse_clean_output(text):
    """Parse JSON output from the model. Returns dict or None."""
    # Strip thinking blocks if present
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()
    text = re.sub(r"<think>.*", "", text, flags=re.DOTALL).strip()

    # Try to extract JSON block
    json_match = re.search(r"\{[\s\S]*\}", text)
    if json_match:
        try:
            result = json.loads(json_match.group())
            # Validate required fields
            required = ["question", "answer", "is_correct"]
            if all(k in result for k in required):
                # Ensure corrected_answer exists
                if "corrected_answer" not in result:
                    result["corrected_answer"] = result["answer"]
                if "verification_note" not in result:
                    result["verification_note"] = ""
                return result
        except json.JSONDecodeError:
            pass

    return None


def load_dataset(path):
    """Load dataset from .csv or .xlsx file. Returns list of dicts with 'question' and 'answer'."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {h: (str(v).strip() if v is not None else "") for h, v in zip(headers, row)}
            if d.get("question"):
                rows.append(d)
        return rows
    else:
        rows = []
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
        return rows


def main(args):
    print(f"Loading dataset from {args.input}")
    rows = load_dataset(args.input)
    print(f"Total rows: {len(rows)}")

    questions = [r["question"] for r in rows]
    answers = [r["answer"] for r in rows]

    # Load model
    print(f"Loading model: {args.model} (TP={args.tp})")
    tokenizer = AutoTokenizer.from_pretrained(args.model, trust_remote_code=True)
    llm = LLM(
        model=args.model,
        tensor_parallel_size=args.tp,
        trust_remote_code=True,
        max_model_len=args.max_model_len,
        gpu_memory_utilization=args.gpu_memory_utilization,
        enforce_eager=True,
    )

    sampling_params = SamplingParams(
        temperature=0.0,
        max_tokens=args.max_tokens,
        n=1,
    )

    # Prepare output
    os.makedirs(args.output_dir, exist_ok=True)
    out_csv = os.path.join(args.output_dir, "math_cleaned.csv")
    out_log = os.path.join(args.output_dir, "math_clean_log.jsonl")
    open(out_log, "w").close()

    cleaned_rows = []
    total_correct = 0
    total_corrected = 0
    total_parse_fail = 0

    total_batches = (len(questions) + args.batch_size - 1) // args.batch_size

    for batch_idx in range(total_batches):
        start = batch_idx * args.batch_size
        end = min(start + args.batch_size, len(questions))
        batch_q = questions[start:end]
        batch_a = answers[start:end]

        prompts = build_clean_prompts(batch_q, batch_a, tokenizer)

        # Retry with smaller sub-batches on failure
        outputs = None
        cur_batch_size = len(prompts)
        for attempt in range(3):
            try:
                t0 = time.time()
                if cur_batch_size >= len(prompts):
                    outputs = llm.generate(prompts, sampling_params)
                else:
                    # Split into sub-batches
                    outputs = []
                    for sub_start in range(0, len(prompts), cur_batch_size):
                        sub_prompts = prompts[sub_start:sub_start + cur_batch_size]
                        sub_out = llm.generate(sub_prompts, sampling_params)
                        outputs.extend(sub_out)
                elapsed = time.time() - t0
                break
            except Exception as e:
                cur_batch_size = max(1, cur_batch_size // 2)
                print(f"  Batch {batch_idx + 1} attempt {attempt + 1} failed: {e}")
                print(f"  Retrying with sub-batch size {cur_batch_size}...")
                if attempt == 2:
                    print(f"  FATAL: Batch {batch_idx + 1} failed after 3 attempts, keeping originals")
                    elapsed = 0.0

        f_log = open(out_log, "a", encoding="utf-8")

        for i in range(len(batch_q)):
            idx = start + i
            log_entry = {
                "index": idx,
                "original_question": batch_q[i],
                "original_answer": batch_a[i],
            }

            if outputs is not None and i < len(outputs):
                text = outputs[i].outputs[0].text
                result = parse_clean_output(text)
                log_entry["raw_output"] = text
            else:
                result = None
                log_entry["raw_output"] = ""

            if result is not None:
                final_answer = result["corrected_answer"] if not result["is_correct"] else result["answer"]
                cleaned_rows.append({
                    "question": result["question"],
                    "answer": final_answer,
                })
                log_entry.update(result)

                if result["is_correct"]:
                    total_correct += 1
                else:
                    total_corrected += 1
            else:
                # Parse failed or batch failed — keep original
                cleaned_rows.append({
                    "question": batch_q[i],
                    "answer": batch_a[i],
                })
                log_entry["parse_error"] = True
                total_parse_fail += 1

            f_log.write(json.dumps(log_entry, ensure_ascii=False) + "\n")

        f_log.close()

        total_done = len(cleaned_rows)
        print(
            f"Batch {batch_idx + 1}/{total_batches}: {end}/{len(questions)} done "
            f"({elapsed:.1f}s) | correct: {total_correct}, corrected: {total_corrected}, "
            f"parse_fail: {total_parse_fail}"
        )

    # Write cleaned CSV
    with open(out_csv, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["question", "answer"])
        writer.writeheader()
        writer.writerows(cleaned_rows)

    # Summary
    total = len(cleaned_rows)
    print(f"\n{'='*60}")
    print(f"Results:")
    print(f"  Total:         {total}")
    print(f"  Correct:       {total_correct} ({total_correct/total*100:.1f}%)")
    print(f"  Corrected:     {total_corrected} ({total_corrected/total*100:.1f}%)")
    print(f"  Parse failed:  {total_parse_fail} ({total_parse_fail/total*100:.1f}%)")
    print(f"{'='*60}")
    print(f"Cleaned CSV:  {out_csv}")
    print(f"Detailed log: {out_log}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Clean math dataset with 72B model")
    parser.add_argument("--model", type=str, default="Qwen/Qwen2.5-72B-Instruct")
    parser.add_argument("--input", type=str, required=True)
    parser.add_argument("--output_dir", type=str, default="../cleaned")
    parser.add_argument("--tp", type=int, default=4)
    parser.add_argument("--max_model_len", type=int, default=8192)
    parser.add_argument("--gpu_memory_utilization", type=float, default=0.9)
    parser.add_argument("--max_tokens", type=int, default=2048,
                        help="Max generation tokens per sample")
    parser.add_argument("--batch_size", type=int, default=32,
                        help="Number of rows per batch")
    args = parser.parse_args()
    main(args)
