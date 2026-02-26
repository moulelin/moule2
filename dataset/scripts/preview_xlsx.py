"""Preview math.xlsx content to check for encoding issues."""
import openpyxl
import sys

path = sys.argv[1] if len(sys.argv) > 1 else "/home/x-qlan1/code/moule2/dataset/raw/math.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active

print(f"Headers: {[cell.value for cell in ws[1]]}")
print(f"Total rows: {ws.max_row - 1}\n")

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
    q = str(row[0]).strip() if row[0] else ""
    a = str(row[1]).strip() if row[1] else ""
    print(f"[{i}] Q: {q}")
    print(f"    A: {a}")
    print()
